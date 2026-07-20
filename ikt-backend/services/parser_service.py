import os
import csv
from .ocr_service import perform_ocr

# Try imports with graceful fallbacks
try:
    import fitz  # PyMuPDF
except ImportError:
    fitz = None

try:
    import docx
except ImportError:
    docx = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

def parse_pdf(file_path: str) -> str:
    if not fitz:
        return "PyMuPDF not installed. Unable to parse PDF."
    text = ""
    try:
        doc = fitz.open(file_path)
        for page in doc:
            text += page.get_text() + "\n"
        doc.close()
    except Exception as e:
        print(f"Error parsing PDF: {e}")
    return text.strip()

def parse_docx(file_path: str) -> str:
    if not docx:
        return "python-docx not installed. Unable to parse Word document."
    text = ""
    try:
        doc = docx.Document(file_path)
        for para in doc.paragraphs:
            text += para.text + "\n"
    except Exception as e:
        print(f"Error parsing DOCX: {e}")
    return text.strip()

def parse_xlsx(file_path: str) -> str:
    if not openpyxl:
        return "openpyxl not installed. Unable to parse Excel sheet."
    text = ""
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            text += f"\n--- Sheet: {sheet} ---\n"
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                row_str = ", ".join([str(val) if val is not None else "" for val in row])
                if row_str.strip():
                    text += row_str + "\n"
    except Exception as e:
        print(f"Error parsing XLSX: {e}")
    return text.strip()

def parse_csv(file_path: str) -> str:
    text = ""
    try:
        with open(file_path, mode='r', encoding='utf-8', errors='ignore') as f:
            reader = csv.reader(f)
            for row in reader:
                row_str = ", ".join(row)
                text += row_str + "\n"
    except Exception as e:
        print(f"Error parsing CSV: {e}")
    return text.strip()

def parse_document(file_path: str, filename: str) -> str:
    ext = filename.split(".")[-1].lower()
    if ext == "pdf":
        return parse_pdf(file_path)
    elif ext == "docx":
        return parse_docx(file_path)
    elif ext == "xlsx":
        return parse_xlsx(file_path)
    elif ext in ["csv", "txt"]:
        return parse_csv(file_path)
    elif ext in ["png", "jpg", "jpeg"]:
        return perform_ocr(file_path)
    else:
        return ""
